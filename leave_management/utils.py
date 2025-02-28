import csv
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context

def export_to_excel(data, filename, headers):
    """
    Generic function to export data to Excel
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row, item in enumerate(data, 2):
        for col, (key, value) in enumerate(item.items(), 1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save(response)
    return response

def export_to_pdf(template_src, context_dict, filename):
    """
    Generic function to export data to PDF using a template
    """
    template = get_template(template_src)
    html = template.render(context_dict)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    # Create PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    return response 